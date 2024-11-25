#!/usr/bin/env python3

#disable cache usage in the Python so __pycache__ isn't formed. If you don't do this using 'nextflow run cdcgov/phoenix...' a second time will causes and error
import sys
sys.dont_write_bytecode = True
import glob
import pandas as pd
import numpy as np
import argparse
import xlsxwriter as ws
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import load_workbook
from itertools import chain
from GRiPHin import order_ar_gene_columns, Combine_dfs, big5_check, write_to_excel, convert_excel_to_tsv, sort_qc_through_spec2_dataframe

##Makes a summary Excel file when given a series of output summary line files from PhoeNiX
##Usage: >python GRiPHin.py -g1 ./samplesheet.csv -a ResGANNCBI_20220915_srst2.fasta -c control_file.csv -o output --phoenix --scaffolds
## Written by Jill Hagey (qpk9@cdc.gov)

# Function to get the script version
def get_version():
    return "1.0.0"

def parseArgs(args=None):
    parser = argparse.ArgumentParser(description='''Script to create new griphin excel sheet by combining two griphin summaries. The -g2 is considered the "new" file and thus when 
    samples with data in the -g1 file will have values overwritten to be the values in the -g2 file.''')
    parser.add_argument('-g1', '--old_griphin', default=None, required=False, dest='griphin_old', help='The first griphin excel file to combine.')
    parser.add_argument('-g2', '--new_griphin', required=False, dest='griphin_new', help='The second griphin excel file to combine.')
    parser.add_argument('-o', '--output', required=False, default=None,dest='output', help='Name of output file default is GRiPHin_Summary.xlsx.')
    parser.add_argument('--griphin_list', required=False, action='store_true',default=False, dest='griphin_list', help='pass instead of -g1/-g2 when you want to combine more than 2 griphins.')
    parser.add_argument('--coverage', default=30, required=False, dest='set_coverage', help='The coverage cut off default is 30x.')
    parser.add_argument('--version', action='version', version=get_version())# Add an argument to display the version
    return parser.parse_args()

#set colors for warnings so they are seen
CRED = '\033[91m'+'\nWarning: '
CEND = '\033[0m'

def read_excels(file_path1, file_path2):
    pd.set_option('display.max_colwidth', None)
    #get number of footer lines for each file to skip when reading in as a pd
    footer_lines1 = detect_footer_lines(file_path1)
    footer_lines2 = detect_footer_lines(file_path2)
    # Read the Excel file, skipping the first row and using the second row as the header
    try: #check that this is an excel file
        df_1 = pd.read_excel(file_path1,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines1,engine='openpyxl')
        df_1.insert(0, 'UNI', df_1.apply(lambda x:'%s/%s/%s' % (x['Parent_Folder'],x['Data_Location'],x['WGS_ID']),axis=1))
        df_1.set_index('UNI')
        ###!print("JJJ0:0A", df_1['UNI'])
        ###!print("JJJ0:0B", df_1.columns.tolist())
        ###!print("JJJ0:0C", df_1.index)
    except Exception as e:
        raise ValueError(f"The input file is not a valid Excel file: {file_path1}")
    try: #check that this is an excel file
        df_2 = pd.read_excel(file_path2,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines2,engine='openpyxl')
        df_2.insert(0, 'UNI', df_2.apply(lambda x:'%s/%s/%s' % (x['Parent_Folder'],x['Data_Location'],x['WGS_ID']),axis=1))
        df_2.set_index('UNI')
        ###!print("JJJ0:1A", df_2['UNI'])
        ###!print("JJJ0:1B", df_2.columns.tolist())
        ###!print("JJJ0:0C", df_2.index)
    except Exception as e:
        raise ValueError(f"The input file is not a valid Excel file: {file_path2}")
    # check that we have the files from the same entry
    ###!print("JJJ1A:",df_1.columns.tolist())
    ###!print("JJJ1C:", file_path1)
    ###!print("JJJ2A:",df_2.columns.tolist())
    ###!print("JJJ2C:", file_path2)
    phoenix, shiga, centar_1, centar_2, all_centar = check_column_presence(df_1, file_path1, df_2, file_path2)
    ###!print("JJJ0:",phoenix, shiga, centar_1, centar_2, all_centar)
    ###!print("JJJ1B:",df_1.columns.tolist())
    ###!print("JJJ2B:",df_2.columns.tolist())
    #parse old griphin df
    if centar_1 == True:
        ordered_centar_df, centar_df_lens,centar_df_column_names = split_centar_df(centar_1, file_path1)
        ###!print("FILE1:", file_path1, "has CENTAR")
        df1_qc, df1_gene_hits_with_centar = split_dataframe(df_1,"Toxinotype")
        ###!print("JJJ1:0A", df1_qc.index)
        ###!print("JJJ1:0B", df1_qc.columns.tolist())
        df1_centar, df1_gene_hits = split_dataframe(df1_gene_hits_with_centar,"AR_Database")
        ###!print("JJJ1:0C", df1_centar.columns.tolist())
        df1_ar, df1_pf_hv = split_dataframe(df1_gene_hits,"HV_Database")
        ###!print("JJJ1:0D", df1_ar.columns.tolist())
        df1_hv, df1_pf = split_dataframe(df1_pf_hv,"Plasmid_Replicon_Database")
        ###!print("JJJ1:0E", df1_pf.columns.tolist())
    else:
        df1_qc, df1_gene_hits = split_dataframe(df_1,"AR_Database")
        df1_ar, df1_pf_hv = split_dataframe(df1_gene_hits,"HV_Database")
        df1_hv, df1_pf_tar = split_dataframe(df1_pf_hv,"Plasmid_Replicon_Database")
        df1_pf, df1_blanktar = split_dataframe(df1_pf_tar,"Toxinotype")
        ###!print("JJJ61B:", df1_pf.columns.tolist())
        
    #parse new griphin df
    if centar_2 == True:
        ordered_centar_df, centar_df_lens, centar_df_column_names = split_centar_df(centar_2, file_path2)
        ###!print("FILE2:", file_path2, "has CENTAR")
        df2_qc, df2_gene_hits_with_centar = split_dataframe(df_2,"Toxinotype")
        ###!print("JJJ1:2A", df1_qc.index)
        ###!print("JJJ1:2B", df2_qc.columns.tolist())
        df2_centar, df2_gene_hits = split_dataframe(df2_gene_hits_with_centar,"AR_Database")
        ###!print("JJJ1:2C", df2_centar.columns.tolist())
        df2_ar, df2_pf_hv = split_dataframe(df2_gene_hits,"HV_Database")
        ###!print("JJJ1:2D", df2_ar.columns.tolist())
        df2_hv, df2_pf = split_dataframe(df2_pf_hv,"Plasmid_Replicon_Database")
        ###!print("JJJ1:2E", df2_pf.columns.tolist())
        if centar_1:
            ###!print("Combining centar from 1 and 2")
            ###!print("JJJ1D:",df1_centar.columns.tolist())
            ###!print("JJJ2D:",df2_centar.columns.tolist())
            centar_headlines=[['UNI', 'Toxinotype', 'Toxin-A_sub-type', 'tcdA', 'Toxin-B_sub-type', 'tcdB'], ['tcdC_Variant', 'tcdC other mutations', 'tcdC', 'tcdR', 'tcdE', 'cdtA', 'cdtB', 'cdtR_Variant', 'cdtR other mutations', 'cdtR', 'cdtAB1', 'cdtAB2', 'PaLoc_NonTox other mutations'], ['gyrA known mutations', 'gyrA other mutations', 'gyrA', 'gyrB known mutations', 'gyrB other mutations', 'gyrB', 'dacS known mutations', 'dacS other mutations', 'dacS', 'feoB known mutations', 'feoB other mutations', 'feoB', 'fur known mutations', 'fur other mutations', 'fur', 'gdpP known mutations', 'gdpP other mutations', 'gdpP', 'glyC known mutations', 'glyC other mutations', 'glyC', 'hemN known mutations', 'hemN other mutations', 'hemN', 'hsmA known mutations', 'hsmA other mutations', 'hsmA', 'lscR known mutations', 'lscR other mutations', 'lscR', 'marR known mutations', 'marR other mutations', 'marR', 'murG known mutations', 'murG other mutations', 'murG', 'nifJ known mutations', 'nifJ other mutations', 'nifJ', 'PNimB known mutations', 'PNimB other mutations', 'PNimB', 'PNimB |','rpoB known mutations', 'rpoB other mutations', 'rpoB', 'rpoC known mutations', 'rpoC other mutations', 'rpoC', 'sdaB known mutations', 'sdaB other mutations', 'sdaB', 'thiH known mutations', 'thiH other mutations', 'thiH', 'vanR known mutations', 'vanR other mutations', 'vanR', 'vanS known mutations', 'vanS other mutations', 'vanS'], ['CEMB RT Crosswalk', 'Inferred RT', 'Probability', 'ML Note', 'Plasmid Info']]
            found_headlines = []
            found_headlines2 = [[],[],[],[]]
            centar_df_lens=[-1,0,0,0]
            for section in range(0,len(centar_headlines)):
                for liner in centar_headlines[section]:
                    if liner in df1_centar.columns.tolist() or liner in df2_centar.columns.tolist():
                        ###!print("JJJ2D:A", liner)
                        found_headlines.append(liner)
                        found_headlines2[section].append(liner)
                        centar_df_lens[section]+=1
            ###!print("JJJ3D:",found_headlines, centar_df_lens)
            ###!print("JJJ4A:",len(df1_centar.columns.tolist()))
            ###!with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
            ###!    print(df1_centar)
            ###!print("JJJ4B:",len(df2_centar.columns.tolist()))
            ###!with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
            ###!    print(df2_centar)
            ###!print("JJJ4C:",cols_to_use, len(cols_to_use))
            combined_centar_df = pd.concat([df1_centar, df2_centar], axis = 0)
            ###!with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
            ###!    print(combined_centar_df)
            ###!print("JJJ4D:",combined_centar_df.columns.tolist(), len(combined_centar_df.columns.tolist()))
            combined_ordered_centar_df = combined_centar_df[found_headlines]
            centar_df_column_names = found_headlines2
            ###!print("JJJ5A:", centar_df_lens)
            ###!with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
            ###!    print(combined_ordered_centar_df)
            ###!print("JJJ4E:",combined_ordered_centar_df.columns.tolist(), len(combined_ordered_centar_df.columns.tolist()))
        else:
            centar_headlines=[['UNI', 'Toxinotype', 'Toxin-A_sub-type', 'tcdA', 'Toxin-B_sub-type', 'tcdB'], ['tcdC_Variant', 'tcdC other mutations', 'tcdC', 'tcdR', 'tcdE', 'cdtA', 'cdtB', 'cdtR_Variant', 'cdtR other mutations', 'cdtR', 'cdtAB1', 'cdtAB2', 'PaLoc_NonTox other mutations'], ['gyrA known mutations', 'gyrA other mutations', 'gyrA', 'gyrB known mutations', 'gyrB other mutations', 'gyrB', 'dacS known mutations', 'dacS other mutations', 'dacS', 'feoB known mutations', 'feoB other mutations', 'feoB', 'fur known mutations', 'fur other mutations', 'fur', 'gdpP known mutations', 'gdpP other mutations', 'gdpP', 'glyC known mutations', 'glyC other mutations', 'glyC', 'hemN known mutations', 'hemN other mutations', 'hemN', 'hsmA known mutations', 'hsmA other mutations', 'hsmA', 'lscR known mutations', 'lscR other mutations', 'lscR', 'marR known mutations', 'marR other mutations', 'marR', 'murG known mutations', 'murG other mutations', 'murG', 'nifJ known mutations', 'nifJ other mutations', 'nifJ', 'PNimB known mutations', 'PNimB other mutations', 'PNimB', 'PNimB |','rpoB known mutations', 'rpoB other mutations', 'rpoB', 'rpoC known mutations', 'rpoC other mutations', 'rpoC', 'sdaB known mutations', 'sdaB other mutations', 'sdaB', 'thiH known mutations', 'thiH other mutations', 'thiH', 'vanR known mutations', 'vanR other mutations', 'vanR', 'vanS known mutations', 'vanS other mutations', 'vanS'], ['CEMB RT Crosswalk', 'Inferred RT', 'Probability', 'ML Note', 'Plasmid Info']]
            found_headlines = []
            found_headlines2 = [[],[],[],[]]
            centar_df_lens=[-1,0,0,0]
            for section in range(0,len(centar_headlines)):
                for liner in centar_headlines[section]:
                    if liner in df2_centar.columns.tolist():
                        ###!print("JJJ2D:B", liner)
                        found_headlines.append(liner)
                        found_headlines2[section].append(liner)
                        centar_df_lens[section]+=1
            combined_ordered_centar_df = ordered_centar_df
    else:
        df2_qc, df2_gene_hits = split_dataframe(df_2,"AR_Database")
        df2_ar, df2_pf_hv = split_dataframe(df2_gene_hits,"HV_Database")
        df2_hv, df2_pf_tar = split_dataframe(df2_pf_hv,"Plasmid_Replicon_Database")
        df2_pf, df2_blanktar = split_dataframe(df2_pf_tar,"Toxinotype")
        ###!print("JJJ62B:", df2_pf.columns.tolist())
        if centar_1:
            for section in range(0,len(centar_headlines)):
                for liner in centar_headlines[section]:
                    if liner in df1_centar.columns.tolist():
                        ###!print("JJJ2D:B", liner)
                        found_headlines.append(liner)
                        found_headlines2[section].append(liner)
                        centar_df_lens[section]+=1
            combined_ordered_centar_df = ordered_centar_df
            combined_ordered_centar_df = df1_centar
        else:
            combined_ordered_centar_df=pd.DataFrame()
            centar_df_lens=[0,0,0,0]
            centar_df_column_names=[]
    ###!with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
    ###!    print(combined_ordered_centar_df)

    #combine qc columns
    combined_df_qc = combine_qc_dataframes(df1_qc, df2_qc)
    #combine ar columns
    combined_df_ar, samples_to_add = combine_gene_dataframes(df1_ar, df2_ar)
    ###!print("JJJ4", combined_df_ar)
    #make sure the order of the ar genes is correct
    order_combined_ar_df = order_ar_gene_columns(combined_df_ar)
    ###!print("Adding sample(s) to the GRiPHin summary:", samples_to_add.tolist())
    # combine pf and hv dataframes
    ###!print("JJJ5:",df1_pf.columns.tolist())
    ###!print("JJJ6:",df2_pf.columns.tolist())
    combined_df_pf, samples_to_add = combine_gene_dataframes(df1_pf, df2_pf)
    combined_df_hv, samples_to_add = combine_gene_dataframes(df1_hv, df2_hv)
    ###!print("CCC:", ordered_centar_df)
    
    return combined_df_qc, order_combined_ar_df, combined_df_pf, combined_df_hv, phoenix, shiga, all_centar, combined_ordered_centar_df, centar_df_lens, centar_df_column_names

def split_centar_df(centar, excel):
    ###!print("DDD0:", type(excel), excel)
    footer_lines1 = detect_footer_lines(excel)
    df = pd.read_excel(excel,
        header=[0, 1],    # Use the second row as the header
        skipfooter=footer_lines1,engine='openpyxl')
    ###!print("DDD:1", df.columns.tolist())
    df.insert(0, 'UNI', df.apply(lambda x:'%s/%s/%s' % (x[('PHoeNIx Summary','Parent_Folder')],x[('PHoeNIx Summary', 'Data_Location')],x[('PHoeNIx Summary', 'WGS_ID')]),axis=1))
    df.set_index('UNI')
    if centar == True:
        df_1 = pd.read_excel(excel,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines1,engine='openpyxl')
        df_1.insert(0, 'UNI', df_1.apply(lambda x:'%s/%s/%s' % (x['Parent_Folder'],x['Data_Location'],x['WGS_ID']),axis=1))
        df_1.set_index('UNI')
        # Get the column names
        columns = df_1.columns.tolist()
        # Find the indices of the specified columns
        ## Might need to fiddle with MLST Clade here eventually...might not if original griphins get made right
        start_index = columns.index('Toxinotype')
        end_index = columns.index('AR_Database')
        # Get the columns between the specified columns
        columns_between = [columns[0]]+columns[start_index:end_index]
        # Subset the DataFrame
        print("DDD:",columns_between)
        ordered_centar_df = df_1[columns_between]
        print("DDD2:", ordered_centar_df)
        print("DDD3:", df['Toxin A/B Variants'].columns, len(df['Toxin A/B Variants'].columns))
        print("DDD4:", df['Other Toxins'].columns, len(df['Other Toxins'].columns))
        print("DDD5:", df['C. difficile Specific AR Mutations'].columns, len(df['C. difficile Specific AR Mutations'].columns))
        print("DDD6:", df['ML Predicted Ribotype'].columns, len(df['ML Predicted Ribotype'].columns))
        #get length of sub dataframes
        centar_df_lens = [ len(df['Toxin A/B Variants'].columns), len(df['Other Toxins'].columns), len(df['C. difficile Specific AR Mutations'].columns), len(df['ML Predicted Ribotype'].columns) ] # have to have it in a list for sum() later
        centar_df_column_names = [ df['Toxin A/B Variants'].columns, df['Other Toxins'].columns, df['C. difficile Specific AR Mutations'].columns, df['ML Predicted Ribotype'].columns ] # have to have it in a list for sum() later
    else:
        centar_df_lens = [0,0,0,0]
        centar_df_column_names = [[],[],[],[]]
        ordered_centar_df = pd.DataFrame()
    return ordered_centar_df, centar_df_lens, centar_df_column_names

def combine_gene_dataframes(old_df, new_df):
    ###!print("OOO1A:", old_df.columns.tolist())
    ###!print("OOO1B:", new_df.columns.tolist())
    # Ensure the first column is 'WGS_ID'
    if new_df.columns[0] != 'UNI' or old_df.columns[0] != 'UNI':
        raise ValueError("The first column in both dataframes must be 'UNI'")
    # Set UNI as index
    ###!print("OOO1:", old_df.columns.tolist())
    ###!print("OOO2:", new_df.columns.tolist())
    df1_gene = old_df.set_index('UNI')
    df2_gene = new_df.set_index('UNI')
    ###!print("OOO2A:", df1_gene.columns.tolist())
    ###!print("OOO2B:", df2_gene.columns.tolist())
    # Identify samples to add and print them
    samples_to_add = df2_gene.index.difference(df1_gene.index)
    ###!print("OOO2C:", df1_gene.columns.tolist())
    ###!print("OOO2D:", df2_gene.columns.tolist())
    # Combine dataframes, prioritizing new_df values and aligning columns
    combined_df = df1_gene.combine_first(df2_gene)
    ###!print("OOO3:", combined_df.columns.tolist())
    # Add missing columns from new_df to old_df
    columns_to_add = [col for col in df2_gene.columns if col not in df1_gene.columns and col not in ["AR_Database", "HV_Database","Plasmid_Replicon_Database", "WGS_ID", "No_AR_Genes_Found", "No_HVGs_Found", "No_Plasmid_Markers", "UNI"]]
    for col in columns_to_add:
        combined_df[col] = combined_df[col].fillna(df2_gene[col])
    ###!print("OOO4:", combined_df.columns.tolist())
    ###!print("OOO4A:", combined_df.index)
    #combined_df['UNI'] = combined_df.index
    # Revert index to column
    combined_df.reset_index(inplace=True)
    ###!print("OOO5:", combined_df.columns.tolist())
    return combined_df, samples_to_add

def combine_qc_dataframes(df1_qc, df2_qc):
    # Convert the 'WGS_ID' columns to the index to facilitate updating
    df1 = df1_qc.set_index('UNI')
    df2 = df2_qc.set_index('UNI')
    # Update df1_qc with matching rows from df2_qc
    df1.update(df2)
    # Append non-matching rows from df2_qc to df1_qc using pd.concat
    combined_df = pd.concat([df1, df2[~df2.index.isin(df1.index)]])
    # Reset the index to restore the 'WGS_ID' column
    combined_df.reset_index(inplace=True)
    return combined_df

def add_blank_centar_columns(df, reference_df):
    # Get the list of column names from the reference dataframe
    columns = reference_df.columns.tolist()

    # Find the indices of the "Toxinotype" and "AR_Database" columns
    try:
        start_index = columns.index("Toxinotype")
        end_index = columns.index("AR_Database")
    except ValueError:
        raise ValueError("One of the required columns ('Toxinotype' or 'AR_Database') is missing in the reference dataframe.")

    # Get all the columns between "Toxinotype" (inclusive) and "AR_Database" (exclusive)
    centar_columns = columns[start_index:end_index]
    # Add blank columns for each centar-related column
    for col in centar_columns:
        if col not in df.columns:
            df[col] = pd.NA  # Fill with NaNs or blank
    return df

def check_column_presence(df1, df1_path, df2, df2_path):
    df1_has_column = "BUSCO_Lineage" in df1.columns
    df2_has_column = "BUSCO_Lineage" in df2.columns
    if df1_has_column and df2_has_column: # set that its a CDC PHOENIX run
        phoenix = False
    elif not df1_has_column and not df2_has_column: # set that its a PHOENIX run
        phoenix = True
    else:
        if df1_has_column:
            raise ValueError(f"{CRED}The old griphin file was produced from -entry CDC_PHOENIX and the new griphin summary wasn't. These files aren't compatible.{CEND}")
        else:
            raise ValueError(f"{CRED}The new griphin file was produced from -entry CDC_PHOENIX and the old griphin summary wasn't. These files aren't compatible.{CEND}")
    #check for centar
    ###!print("NNN1:", df1.columns.tolist())
    ###!print("NNN2:", df2.columns.tolist())
    if "Toxin-A_sub-type" in df1.columns or "Toxin-A_sub-type" in df2.columns:
        if "Toxin-A_sub-type" not in df1.columns:
            print("Adding centar blank columns to " + df1_path)
            df1 = add_blank_centar_columns(df1, df2)
            centar_1 = False
        else:
            centar_1 = True
        if "Toxin-A_sub-type" not in df2.columns:
            print("Adding centar blank columns to " + df2_path)
            df2 = add_blank_centar_columns(df2, df1)
            centar_2 = False
        else:
            centar_2 = True
    else:
        centar_1 = centar_2 = False
    if centar_1 == True or centar_2 == True:
        all_centar = True
    else:
        all_centar = False
    #check for shigapass
    if "ShigaPass_Organism" in df1.columns or "ShigaPass_Organism" in df2.columns:
        shiga = True
    else:
        shiga = False
    return phoenix, shiga, centar_1, centar_2, all_centar

def split_dataframe(df, split_column):
    # Ensure the first column is 'UNI'
    if df.columns[0] != 'UNI':
        raise ValueError("The first column must be 'UNI'")
    # Find the index of the split_column
    split_index = df.columns.get_loc(split_column)
    ###!print("UUU0:", df.columns.tolist())
    # Create two DataFrames based on the split index, including 'UNI' at the beginning
    df_before = df.iloc[:, [0] + list(range(1, split_index))]  # All columns before the split column + 'UNI'
    #df_before = df.iloc[:, list(range(1, split_index))]  # All columns before the split column + 'UNI'
    df_after = df.iloc[:, [0] + list(range(split_index, df.shape[1]))]  # The split column and all columns after it + 'UNI'
    ###!print("UUU1:", df_before.columns.tolist())
    ###!print("UUU2:", df_after.columns.tolist())
    return df_before, df_after

def add_and_concatenate(df1, df2):
    # Step 1: Preserve column order from df1 and append missing columns from df2
    all_columns = list(df1.columns) + [col for col in df2.columns if col not in df1.columns]
    # Step 1: Find the union of columns in both DataFrames
    #all_columns = set(df1.columns).union(set(df2.columns))
    # Step 2: Add missing columns to each DataFrame at once by reindexing
    df1 = df1.reindex(columns=all_columns)
    df2 = df2.reindex(columns=all_columns)
    # Step 3: Concatenate the two DataFrames row-wise
    result = pd.concat([df1, df2], ignore_index=True, sort=False)
    # Step 4: Reset the index and return the final DataFrame
    result.reset_index(drop=True, inplace=True)
    return result

def detect_footer_lines(file_path, sheet_name=0):
    # Load the workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name if isinstance(sheet_name, str) else wb.sheetnames[sheet_name]]
    footer_lines = 3 #set to 3 because review by row has data in another cell and there are two all blank spacing rows
    is_footer_section = False
    # Count footer lines
    for row in reversed(tuple(ws.rows)):
        first_cell = row[0].value
        other_cells = [cell.value for cell in row[1:]]
        # Check if the first cell has text and all other cells are empty
        if isinstance(first_cell, str) and all(cell is None for cell in other_cells):
            footer_lines += 1
            is_footer_section = True
        elif is_footer_section:
            break
    return footer_lines

def read_excel(file_path, old_phoenix, reference_qc_df, reference_centar_df):
    #get number of footer lines for each file to skip when reading in as a pd
    footer_lines = detect_footer_lines(file_path)
    # Read the Excel file, skipping the first row and using the second row as the header
    try: #check that this is an excel file
        df = pd.read_excel(file_path,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines,engine='openpyxl')
        df.insert(0, 'UNI', df.apply(lambda x:'%s/%s/%s' % (x['Parent_Folder'],x['Data_Location'],x['WGS_ID']),axis=1))
        df.set_index('UNI')
    except Exception as e:
        raise ValueError(f"The input file is not a valid Excel file: {file_path}")
    # check that we have the files from the same entry
    if "BUSCO_Lineage" in df.columns: # set that its a CDC PHOENIX run
        phoenix = False
    else:
        phoenix = False
    if old_phoenix != phoenix: 
        raise ValueError(f"{CRED}The one griphin file in your set was produced from -entry CDC_PHOENIX and some were not wasn't. These files aren't compatible to combine. {CEND}")
    #check for centar
    if "Toxin-A_sub-type" not in df.columns:
        print("Adding centar blank columns to " + file_path)
        for col in reference_centar_df.columns:
            if col not in df.columns:
                df[col] = pd.NA  # Fill with NaNs or blank
        centar = all_centar = False
    else:
        centar = all_centar = True
    #check for shigapass
    if "ShigaPass_Organism" in df.columns:
        shiga = True
    else:
        shiga = False
    #parse old griphin df
    ###!print("RRR-2:", file_path)
    ###!print("RRR-1:", centar)
    ordered_centar_df, centar_df_lens, centar_df_column_names = split_centar_df(centar, file_path)
    #centar_df_lens[0] += 1
    #centar_df_lens = [x + 1 for x in centar_df_lens]
    ###!print("RRR0:", ordered_centar_df.columns.tolist())
    ###!print("RRR1:", centar_df_lens)
    ###!print("RRR2:", centar_df_column_names)
    if centar == True:
        df_qc, df_gene_hits_with_centar = split_dataframe(df,"Toxinotype")
        ###!print("RRR3:")
        ###!print(df_qc["UNI"])
        df_centar, df_gene_hits = split_dataframe(df_gene_hits_with_centar,"AR_Database")
        ###!print("RRR4:")
        ###!print(df_centar["UNI"])
        ordered_centar_df = df_centar #.set_index('WGS_ID')
        ###!print("RRR5:", ordered_centar_df.columns.tolist())
        df_ar, df_pf_hv = split_dataframe(df_gene_hits,"HV_Database")
        df_hv, df_pf = split_dataframe(df_pf_hv,"Plasmid_Replicon_Database")
    else:
        df_qc, df_gene_hits = split_dataframe(df,"AR_Database")
        df_ar, df_pf_hv = split_dataframe(df_gene_hits,"HV_Database")
        df_hv, df_pf = split_dataframe(df_pf_hv,"Plasmid_Replicon_Database")
        # Ensure the first column is 'UNI'
    if df_qc.columns[0] != 'UNI' or reference_qc_df.columns[0] != 'UNI':
        raise ValueError("The first column in both dataframes must be 'UNI'")
    # Set UNI as index
    df1_qc = reference_qc_df.set_index('UNI')
    ref_qc = df_qc.set_index('UNI')
    # Identify samples to add and print them
    samples_to_add = ref_qc.index.difference(df1_qc.index)
    #make sure the order of the ar genes is correct
    order_ar_df = order_ar_gene_columns(df_ar)
    print("Adding sample(s) to the GRiPHin summary:", samples_to_add.tolist())
    ###!print("RRR6:")
    ###!print(ordered_centar_df["UNI"])
    return df_qc, order_ar_df, df_pf, df_hv, phoenix, shiga, centar, ordered_centar_df, centar_df_lens, centar_df_column_names

def update_centar_columns(centar_df_column_names_final, centar_df_column_names):
    ###!print("6660:", centar_df_column_names_final, centar_df_column_names)
    # Iterate over each corresponding pair of lists in centar_df_lens_final and centar_df_lens
    for i, (final_list, new_list) in enumerate(zip(centar_df_column_names_final, centar_df_column_names)):
        # Convert both lists to sets to get unique column names
        unique_columns = set(final_list).union(set(new_list))
        # Update centar_df_lens_final with the unique columns
        centar_df_column_names_final[i] = list(unique_columns)
    # After processing all lists, calculate the total number of unique names in each list
    total_unique_columns = [len(columns) for columns in centar_df_column_names_final]
    ###!print("6661",total_unique_columns)
    centar_headlines=[['UNI', 'Toxinotype', 'Toxin-A_sub-type', 'tcdA', 'Toxin-B_sub-type', 'tcdB'], ['tcdC_Variant', 'tcdC other mutations', 'tcdC', 'tcdR', 'tcdE', 'cdtA', 'cdtB', 'cdtR_Variant', 'cdtR other mutations', 'cdtR', 'cdtAB1', 'cdtAB2', 'PaLoc_NonTox_Variant', 'PaLoc_NonTox other mutations', 'PaLoc_NonTox'], ['gyrA known mutations', 'gyrA other mutations', 'gyrA', 'gyrB known mutations', 'gyrB other mutations', 'gyrB', 'dacS known mutations', 'dacS other mutations', 'dacS', 'feoB known mutations', 'feoB other mutations', 'feoB', 'fur known mutations', 'fur other mutations', 'fur', 'gdpP known mutations', 'gdpP other mutations', 'gdpP', 'glyC known mutations', 'glyC other mutations', 'glyC', 'hemN known mutations', 'hemN other mutations', 'hemN', 'hsmA known mutations', 'hsmA other mutations', 'hsmA', 'lscR known mutations', 'lscR other mutations', 'lscR', 'marR known mutations', 'marR other mutations', 'marR', 'murG known mutations', 'murG other mutations', 'murG', 'nifJ known mutations', 'nifJ other mutations', 'nifJ', 'PNimB known mutations', 'PNimB other mutations', 'PNimB', 'PNimB |','rpoB known mutations', 'rpoB other mutations', 'rpoB', 'rpoC known mutations', 'rpoC other mutations', 'rpoC', 'sdaB known mutations', 'sdaB other mutations', 'sdaB', 'thiH known mutations', 'thiH other mutations', 'thiH', 'vanR known mutations', 'vanR other mutations', 'vanR', 'vanS known mutations', 'vanS other mutations', 'vanS'], ['CEMB RT Crosswalk', 'Inferred RT', 'Probability', 'ML Note', 'Plasmid Info']]
    found_headlines = [ [] for _ in range(0,len(centar_df_column_names_final)) ]
    for index in range(0, len(centar_df_column_names_final)):
        for liner in centar_headlines[index]:
            for col in centar_df_column_names_final[index]:
                ###!print("666X:", index, col,  liner)
                if col == liner:
                    ###!print("6662: Found headline", index, liner)
                    found_headlines[index].append(liner)
    total_unique_columns = [len(found_headlines[0])-1, len(found_headlines[1]), len(found_headlines[2]), len(found_headlines[3])]
    ###!print("666Y:", total_unique_columns)
    return found_headlines, total_unique_columns
    

def main():
    # looping through excel files
    args = parseArgs()
    #figure out the name of the output file 
    if args.output != None:
        output_file = args.output
    else:
        # Derive output file name from input file name
        output_file = args.griphin_new.replace("_GRiPHin_Summary.xlsx", "")
    # checking what the input type is
    if args.griphin_list != False:
        griphin_files = glob.glob("*_GRiPHin_Summary.xlsx")
        if len(griphin_files) < 2:
            raise ValueError(f"{CRED}Need at least two GRiPHin files for combination when using --griphin_list.{CEND}")
        
        # Initialize the combination with the first file
        base_file = griphin_files.pop(0)
        ###!print("QQQ-1:", base_file, griphin_files[0])
        combined_df_qc, combined_df_ar, combined_df_pf, combined_df_hv, phoenix_final, shiga_final, centar_final, ordered_centar_df, centar_df_lens_final, centar_df_column_names_final = read_excels(base_file, griphin_files[0])
        combined_dataframes_1 = [ combined_df_qc, combined_df_ar, combined_df_pf, combined_df_hv, ordered_centar_df ]
        # Iterate over remaining files, progressively combining them with the base
        ###!print("QQQ0:", centar_df_lens_final, centar_df_column_names_final)
        for count, next_file in enumerate(griphin_files[1:], start=1):
            # checking
            ###!print("QQQ1:", next_file)
            combined_df_qc, combined_df_ar, combined_df_pf, combined_df_hv, phoenix, shiga, centar, ordered_centar_df, centar_df_lens_final, centar_df_column_names = read_excel(next_file, phoenix_final, combined_df_qc, ordered_centar_df)
            ###!print("QQQ2:", centar_df_lens_final, centar_df_column_names_final)
            combined_dataframes = [ combined_df_qc, combined_df_ar, combined_df_pf, combined_df_hv, ordered_centar_df ]
            # Update flags
            phoenix_final = phoenix_final or phoenix
            shiga_final = shiga_final or shiga
            centar_final = centar_final or centar
            if count == 1:
                # Iterate over both tuples and combine corresponding DataFrames
                combined_dataframes_final = [ add_and_concatenate(df1, df2) for df1, df2 in zip(combined_dataframes_1, combined_dataframes) ]
                if centar_final == True:
                    # Update centar_df_lens1 with the maximum of each corresponding number
                    #centar_df_lens_final = [max(val1, val2) for val1, val2 in zip(centar_df_lens_final, centar_df_lens)]
                    ###!print("7770:")
                    centar_df_column_names_final, centar_df_lens_final = update_centar_columns(centar_df_column_names_final, centar_df_column_names)
                    ###!print("7770B:", centar_df_column_names_final)
            else:
                combined_dataframes_final = [ add_and_concatenate(df1, df2) for df1, df2 in zip(combined_dataframes_final, combined_dataframes) ]
                if centar_final == True:
                    ###!print("7771:")
                    # Update centar_df_lens1 with the maximum of each corresponding number
                    #centar_df_lens_final = [max(val1, val2) for val1, val2 in zip(centar_df_lens_final, centar_df_lens)]
                    centar_df_column_names_final, centar_df_lens_final = update_centar_columns(centar_df_column_names_final, centar_df_column_names)
                    ###!print("7771B:", centar_df_lens_final, centar_df_column_names_final)
            # Unpack the tuple into individual DataFrames
            for frame in combined_dataframes_final:
                ###!print("88883:", frame)
                frame = sort_qc_through_spec2_dataframe(frame, False)
            combined_df_qc, combined_df_ar, combined_df_pf, combined_df_hv, ordered_centar_df = combined_dataframes_final
    else:
        combined_df_qc, combined_df_ar, combined_df_pf, combined_df_hv, phoenix_final, shiga_final, centar_final, ordered_centar_df, centar_df_lens_final, centar_df_column_names_final = read_excels(args.griphin_new, args.griphin_old)
    ###!print(centar_df_lens_final)
    ###!print(centar_df_column_names_final)
    ###!print(ordered_centar_df)
    #ordered_centar_df = ordered_centar_df.drop('WGS_ID', axis=1)
    # call function from griphin script to combine all dfs
    ###!print("TTT0:", len(combined_df_qc['UNI']), len(combined_df_ar['UNI']), len(combined_df_pf['UNI']), len(combined_df_hv['UNI']))
    final_df, ar_max_col, columns_to_highlight, final_ar_df, pf_db, ar_db, hv_db = Combine_dfs(combined_df_qc, combined_df_ar, combined_df_pf, combined_df_hv, pd.DataFrame(), phoenix_final, centar_final, ordered_centar_df)
    ###!print(len(combined_df_qc.columns.tolist()), combined_df_qc.columns.tolist())
    ###!print("TTT1:", final_df['UNI'])
    #get other information for excel writing
    ###!print("BBBB1:", final_df.columns.tolist())
    combined_df_qc = combined_df_qc.drop('UNI', axis = 1)
    (qc_max_row, qc_max_col) = combined_df_qc.shape
    # Mst account for UNI removal later, so decreasing size by one here before i forget
    qc_max_row -= 1
    ###!print("BBB1B:", qc_max_row,qc_max_col)
    pf_max_col = combined_df_pf.shape[1] - 1 #remove one for the UNI column
    hv_max_col = combined_df_hv.shape[1] - 1 #remove one for the UNI column
    #write excel sheet
    final_df = final_df.drop('UNI', axis=1)
    ar_start_index = qc_max_col + sum(centar_df_lens_final)
    write_to_excel(args.set_coverage, output_file, final_df, qc_max_col, ar_max_col, pf_max_col, hv_max_col, columns_to_highlight, final_ar_df, pf_db, ar_db, hv_db, phoenix_final, shiga_final, centar_final, centar_df_lens_final, ordered_centar_df, ar_start_index)
    #write tsv from excel
    convert_excel_to_tsv(output_file)

if __name__ == '__main__':
    main()